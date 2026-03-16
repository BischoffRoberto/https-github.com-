from fastapi import FastAPI, Request, Form, HTTPException, UploadFile, File
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import pandas as pd
from datetime import datetime
import io
import os
import zipfile
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from typing import Dict, Any, Optional, List, Tuple
import unicodedata

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# Carpeta para listas de usuarios
LISTAS_DIR = "listas"
os.makedirs(LISTAS_DIR, exist_ok=True)

# Archivo para usuarios persistentes
USERS_FILE = "usuarios.json"

# Údtiles para normalizar texto
def _norm_str(s: str) -> str:
    """Normaliza texto quitando espacios, minúsculas y acentos."""
    s2 = str(s).strip()
    s2 = unicodedata.normalize('NFKD', s2).encode('ascii', 'ignore').decode('ascii')
    return s2.lower()

def _strip_leading_zeros(s: str) -> str:
    """Para búsquedas: quita ceros a la izquierda de strings numéricos."""
    s_str = str(s).strip()
    # Si es solo números, quita ceros
    if s_str.isdigit():
        s_str = s_str.lstrip('0') or '0'
    return s_str

def _strip_trailing_dot_zero(s: str) -> str:
    """Quita el sufijo .0 que aparece al leer números desde Excel."""
    s_str = str(s).strip()
    if s_str.endswith('.0'):
        return s_str[:-2]
    return s_str

def _normalize_columns(cols):
    """Normaliza nombres de columnas en un DataFrame."""
    return [_norm_str(c) for c in cols]

def guardar_usuarios():
    """Almacena el diccionario `usuarios` en un JSON."""
    try:
        with open(USERS_FILE, 'w', encoding='utf-8') as f:
            json.dump(usuarios, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error guardando usuarios: {e}")

def cargar_usuarios():
    """Carga el archivo de usuarios si existe, dejando los valores por defecto en caso contrario."""
    global usuarios
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, dict):
                    usuarios = data
        except Exception as e:
            print(f"Error cargando usuarios: {e}")

def guardar_listas():
    """Guarda las listas de usuarios en archivos JSON"""
    for usuario, lista in listas_por_usuario.items():
        if lista:
            archivo = os.path.join(LISTAS_DIR, f"{usuario}.json")
            try:
                with open(archivo, 'w', encoding='utf-8') as f:
                    json.dump(lista, f, ensure_ascii=False, indent=2)
            except Exception as e:
                print(f"Error guardando lista de {usuario}: {e}")

def cargar_listas():
    """Carga las listas de usuarios desde archivos JSON"""
    global listas_por_usuario
    listas_por_usuario = {}
    if not os.path.exists(LISTAS_DIR):
        return
    for archivo in os.listdir(LISTAS_DIR):
        if archivo.endswith('.json'):
            usuario = archivo[:-5]
            ruta = os.path.join(LISTAS_DIR, archivo)
            try:
                with open(ruta, 'r', encoding='utf-8') as f:
                    lista = json.load(f)
                    if isinstance(lista, list):
                        listas_por_usuario[usuario] = lista
            except Exception as e:
                print(f"Error cargando lista de {usuario}: {e}")

def cargar_ean():
    """Carga el archivo EAN.xlsx. IMPORTANTE: Asigna a global ean_df"""
    global ean_df
    try:
        if os.path.exists("EAN.xlsx"):
            temp = pd.read_excel("EAN.xlsx")
            print(f"[EAN] Archivo cargado, columnas originales: {list(temp.columns)}")
            temp.columns = _normalize_columns(temp.columns)
            print(f"[EAN] Columnas normalizadas: {list(temp.columns)}")
            if 'ean' in temp.columns and 'codigo' in temp.columns:
                # Normalizar datos: quitar sufijo .0 y limpiar espacios
                temp['ean'] = temp['ean'].astype(str).apply(_strip_trailing_dot_zero)
                temp['codigo'] = temp['codigo'].astype(str).apply(_strip_trailing_dot_zero)
                if 'descripcion' in temp.columns:
                    temp['descripcion'] = temp['descripcion'].astype(str).str.strip()
                else:
                    temp['descripcion'] = ''
                ean_df = temp
                print(f"[EAN] ✓ EAN.xlsx cargado exitosamente: {len(temp)} filas")
                return
            else:
                print(f"[EAN] ✗ Archivo no tiene columnas EAN o Codigo")
        else:
            print(f"[EAN] EAN.xlsx no existe en cwd={os.getcwd()}")
    except Exception as e:
        print(f"[EAN] ✗ Error cargando EAN.xlsx: {e}")
    ean_df = pd.DataFrame(columns=["ean", "codigo", "descripcion"])
    print(f"[EAN] DataFrame vacío creado")

# --- Inicializar variables globales ---
usuarios: Dict[str, Dict[str, Any]] = {
    "admin": {"password": "admin1234", "ips": [], "activo": True, "rol": "admin", "tlf": "", "legajo": "0001", "nombre": "Administrador", "correo": "", "bloqueado": False},
    "roberto": {"password": "clave123", "ips": [], "activo": True, "rol": "usuario", "tlf": "", "legajo": "0002", "nombre": "Roberto", "correo": "", "bloqueado": False}
}

sesiones: Dict[str, Dict[str, Any]] = {}
listas_por_usuario: Dict[str, List[Dict[str, Any]]] = {}
ean_df: pd.DataFrame = pd.DataFrame(columns=["ean", "codigo"])

# Cargar datos persistentes
try:
    cargar_usuarios()
except Exception as e:
    print(f"Error en cargar_usuarios: {e}")
try:
    cargar_listas()
except Exception as e:
    print(f"Error en cargar_listas: {e}")
try:
    cargar_ean()
except Exception as e:
    print(f"Error en cargar_ean: {e}")

# Inventario base
try:
    df = pd.read_excel("Inventario.xlsx")
    df.columns = _normalize_columns(df.columns)
except FileNotFoundError:
    df = pd.DataFrame(columns=["codigo", "descripcion", "stock"])

# Carpeta para archivos subidos por usuarios
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Carpeta para respaldos de inventario
BACKUP_DIR = "backups"
os.makedirs(BACKUP_DIR, exist_ok=True)

# --- Alertas por correo ---
import smtplib
from email.message import EmailMessage

def send_email_to_user(user: str, subject: str, body: str):
    """Envía correo si SMTP está configurado"""
    smtp_host = os.environ.get('SMTP_HOST')
    smtp_port = int(os.environ.get('SMTP_PORT', '0') or 0)
    smtp_user = os.environ.get('SMTP_USER')
    smtp_pass = os.environ.get('SMTP_PASS')
    user_email = usuarios.get(user, {}).get('correo')
    if not (smtp_host and smtp_port and smtp_user and smtp_pass and user_email):
        return
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = smtp_user
    msg['To'] = user_email
    msg.set_content(body)
    with smtplib.SMTP(smtp_host, smtp_port) as s:
        s.starttls()
        s.login(smtp_user, smtp_pass)
        s.send_message(msg)

def _parse_fecha_vencimiento(fecha_vencimiento: str) -> Optional[date]:
    """Intenta parsear la fecha de vencimiento en varios formatos.

    Retorna un objeto date o None si la fecha no es válida.
    """
    if not fecha_vencimiento:
        return None

    fecha_str = str(fecha_vencimiento).strip()
    formatos = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"]
    for fmt in formatos:
        try:
            return datetime.strptime(fecha_str, fmt).date()
        except ValueError:
            continue

    # Intentar parseo ISO como respaldo
    try:
        return datetime.fromisoformat(fecha_str).date()
    except Exception:
        return None


def estado_vencimiento(fecha_vencimiento: str) -> Tuple[str, int]:
    """Retorna (estado_texto, dias_restantes)"""
    hoy = datetime.today().date()
    fecha = _parse_fecha_vencimiento(fecha_vencimiento)
    if not fecha:
        return ("Fecha inválida", 0)

    dias = (fecha - hoy).days
    if dias < 0:
        return ("Vencido", dias)
    elif dias == 0:
        return ("Se vence hoy", dias)
    elif dias <= 7:
        return ("Crítico (<7 días)", dias)
    return ("Correcto", dias)

# --- Login / Logout ---
@app.get("/")
async def login_page(request: Request):
    ip_actual = request.client.host if request.client else "0.0.0.0"
    return templates.TemplateResponse("login.html", {"request": request, "ip": ip_actual})

@app.post("/login")
async def login(request: Request, usuario: str = Form(...), contrasena: str = Form(...)):
    ip_actual = request.client.host if request.client else "0.0.0.0"
    
    if usuario not in usuarios:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Usuario no encontrado", "ip": ip_actual})
    
    user_data = usuarios[usuario]
    
    if user_data["password"] != contrasena:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Contraseña incorrecta"})
    
    if not user_data["activo"]:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Usuario bloqueado"})
    
    ips = user_data.get("ips", [])
    if ip_actual not in ips:
        ips.append(ip_actual)
        user_data["ips"] = ips

    if usuario != "admin":
        if user_data.get("bloqueado"):
            return templates.TemplateResponse("login.html", {"request": request, "error": "IP bloqueada. Contacta al administrador"})
        if ips and ip_actual not in ips:
            return templates.TemplateResponse("login.html", {"request": request, "error": "Acceso desde IP no autorizada. IP registrada: " + (ips[0] if ips else "ninguna")})
    
    response = RedirectResponse(url="/home", status_code=302)
    # debug output and explicit cookie options
    print(f"[login] setting session_id cookie for {usuario}")
    response.set_cookie("session_id", usuario, httponly=False, samesite="lax", path="/")

    # 👉 Aquí guardamos la sesión
    # guardamos datos de sesión, incluyendo el nombre para uso posterior
    sesiones[usuario] = {"ip": ip_actual, "activo": True, "usuario": usuario}

    try:
        send_email_to_user(usuario, "Acceso al sistema", f"Usuario {usuario} ha iniciado sesión desde IP {ip_actual}.")
    except Exception:
        pass

    return response

@app.get("/home")
async def home(request: Request):
    print(f"[home] incoming cookies: {request.cookies}")
    session_id = request.cookies.get("session_id")
    if not session_id:
        return RedirectResponse(url="/")
    rol = usuarios[session_id]["rol"]
    legajo = usuarios[session_id].get("legajo", "")
    ip_actual = request.client.host if request.client else "0.0.0.0"
    return templates.TemplateResponse("index.html", {"request": request, "usuario": session_id, "rol": rol, "legajo": legajo, "ip": ip_actual})


@app.get("/offline")
async def offline(request: Request):
    return templates.TemplateResponse("offline.html", {"request": request})

@app.get("/logout")
async def logout(request: Request):
    response = RedirectResponse(url="/")
    response.delete_cookie("session_id")
    return response

# --- Panel Admin ---
@app.get("/admin")
async def admin_page(request: Request):
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios[session_id]["rol"] != "admin":
        return RedirectResponse(url="/")
    # obtener IP para mostrar en cabecera
    ip_actual = request.client.host if request.client else "0.0.0.0"
    # preparar preview de inventario (todas filas, limitado si es muy grande)
    try:
        # limitar para no cargar gigas; pero por ahora enviamos todo
        preview_rows = df.to_dict(orient="records")
    except Exception:
        preview_rows = []
    return templates.TemplateResponse("admin.html", {"request": request, "usuarios": usuarios, "ip": ip_actual, "preview": preview_rows})

@app.post("/admin/agregar_usuario")
async def admin_agregar_usuario(request: Request) -> Dict[str, Any]:
    data = await request.json()
    nombre = data.get("usuario")
    contrasena = data.get("contrasena")
    rol = data.get("rol", "usuario")
    nombre_completo = data.get("nombre", "")
    tlf = data.get("tlf", "")
    legajo = data.get("legajo", "")
    correo = data.get("correo", "")
    if not nombre or not contrasena:
        raise HTTPException(status_code=400, detail="Faltan datos")
    if nombre in usuarios:
        raise HTTPException(status_code=400, detail="Usuario ya existe")
    usuarios[nombre] = {"password": contrasena, "ips": [], "activo": True, "rol": rol, "tlf": tlf, "legajo": legajo, "nombre": nombre_completo, "correo": correo, "bloqueado": False}
    guardar_usuarios()
    return {"mensaje": "Usuario agregado", "usuario": {"nombre": nombre, "password": contrasena, "rol": rol, "activo": True, "ips": [], "tlf": tlf, "legajo": legajo, "nombre_completo": nombre_completo, "correo": correo}}

@app.post("/admin/bloquear_usuario")
async def admin_bloquear_usuario(request: Request) -> Dict[str, Any]:
    data = await request.json()
    nombre = data.get("usuario")
    if nombre not in usuarios:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if nombre == "admin":
        raise HTTPException(status_code=400, detail="No se puede bloquear al administrador principal")
    usuarios[nombre]["activo"] = False
    guardar_usuarios()
    return {"mensaje": "Usuario bloqueado", "usuarios": usuarios}

@app.post("/admin/desbloquear_usuario")
async def admin_desbloquear_usuario(request: Request) -> Dict[str, Any]:
    data = await request.json()
    nombre = data.get("usuario")
    if nombre not in usuarios:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if nombre == "admin":
        raise HTTPException(status_code=400, detail="El administrador principal no necesita desbloqueo")
    usuarios[nombre]["activo"] = True
    guardar_usuarios()
    return {"mensaje": "Usuario desbloqueado", "usuarios": usuarios}


@app.post("/admin/modificar_usuario")
async def admin_modificar_usuario(request: Request) -> Dict[str, Any]:
    data = await request.json()
    nombre = data.get("usuario")
    if nombre not in usuarios:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if "contrasena" in data and data["contrasena"]:
        usuarios[nombre]["password"] = data["contrasena"]
    if "rol" in data and data["rol"]:
        usuarios[nombre]["rol"] = data["rol"]
    if "nombre" in data and data["nombre"]:
        usuarios[nombre]["nombre"] = data["nombre"]
    if "tlf" in data and data["tlf"]:
        usuarios[nombre]["tlf"] = data["tlf"]
    if "legajo" in data and data["legajo"]:
        usuarios[nombre]["legajo"] = data["legajo"]
    if "correo" in data and data["correo"]:
        usuarios[nombre]["correo"] = data["correo"]
    guardar_usuarios()
    return {"mensaje": "Usuario modificado"}

@app.post("/admin/bloquear_ip")
async def admin_bloquear_ip(request: Request) -> Dict[str, Any]:
    data = await request.json()
    nombre = data.get("usuario")
    if nombre not in usuarios:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    usuarios[nombre]["bloqueado"] = True
    return {"mensaje": f"Usuario {nombre} bloqueado por IP"}

@app.post("/admin/desbloquear_ip")
async def admin_desbloquear_ip(request: Request) -> Dict[str, Any]:
    data = await request.json()
    nombre = data.get("usuario")
    if nombre not in usuarios:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    usuarios[nombre]["bloqueado"] = False
    usuarios[nombre]["ips"] = []  # Reset IPs para permitir nueva IP
    return {"mensaje": f"Usuario {nombre} desbloqueado. Puede usar nueva IP"}

@app.post("/admin/borrar_usuario")
async def admin_borrar_usuario(request: Request) -> Dict[str, Any]:
    data = await request.json()
    nombre = data.get("usuario")
    if nombre not in usuarios:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if nombre == "admin":
        raise HTTPException(status_code=400, detail="No se puede borrar al administrador principal")
    usuarios.pop(nombre)
    guardar_usuarios()
    return {"mensaje": f"Usuario {nombre} eliminado"}


# --- Archivos subidos por usuarios / administrador ---
@app.post("/admin/upload_archivo")
async def admin_upload_archivo(request: Request, file: UploadFile = File(...)) -> Dict[str, Any]:
    """Sube un archivo Excel como inventario base. Requiere columnas: Código y Descripción."""
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")

    user_dir = os.path.join(UPLOAD_DIR, session_id)
    os.makedirs(user_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    safe_name = (file.filename or "").replace('..', '')
    save_name = f"{timestamp}_{safe_name}"
    path = os.path.join(user_dir, save_name)
    
    # Guardar archivo físicamente
    contents = await file.read()
    with open(path, "wb") as f:
        f.write(contents)
    
    print(f"\n[UPLOAD_INVENTARIO] ===== Inicio =====")
    print(f"[UPLOAD_INVENTARIO] Archivo guardado en: {path}")
    
    global df, listas_por_usuario
    parsed_rows = []
    
    try:
        # Leer archivo
        print(f"[UPLOAD_INVENTARIO] Leyendo archivo...")
        temp = pd.read_excel(path)
        print(f"[UPLOAD_INVENTARIO] Columnas originales: {list(temp.columns)}")
        
        # Normalizar columnas
        temp.columns = _normalize_columns(temp.columns)
        print(f"[UPLOAD_INVENTARIO] Columnas normalizadas: {list(temp.columns)}")
        
        # Validar que existan las columnas requeridas
        if 'codigo' not in temp.columns or 'descripcion' not in temp.columns:
            cols_faltantes = []
            if 'codigo' not in temp.columns:
                cols_faltantes.append('Código')
            if 'descripcion' not in temp.columns:
                cols_faltantes.append('Descripción')
            raise ValueError(f"Faltan columnas: {', '.join(cols_faltantes)}")
        
        print(f"[UPLOAD_INVENTARIO] ✓ Columnas código y descripción encontradas")
        
        # Hacer backup del inventario anterior
        if os.path.exists('Inventario.xlsx'):
            ts = datetime.now().strftime('%Y%m%d%H%M%S')
            backup_name = f"inventario_{ts}.xlsx"
            backup_path = os.path.join(BACKUP_DIR, backup_name)
            try:
                os.replace('Inventario.xlsx', backup_path)
                print(f"[UPLOAD_INVENTARIO] Backup guardado: {backup_name}")
            except Exception as e:
                print(f"[UPLOAD_INVENTARIO] No se pudo hacer backup (no crítico): {e}")
        
        # Guardar como Inventario.xlsx en la raíz
        print(f"[UPLOAD_INVENTARIO] Guardando como Inventario.xlsx...")
        temp.to_excel('Inventario.xlsx', index=False)
        print(f"[UPLOAD_INVENTARIO] ✓ Inventario.xlsx guardado ({len(temp)} filas)")
        
        # Actualizar variable global df
        df = temp
        
        # Preparar preview
        parsed_rows = df.head(50).to_dict(orient='records')
        print(f"[UPLOAD_INVENTARIO] Preview generado: {len(parsed_rows)} filas mostradas")
        # NO tocar las listas existentes de usuarios - solo actualizar el inventario
        print(f"[UPLOAD_INVENTARIO] ===== ✓ ÉXITO =====\n")
        
    except Exception as e:
        print(f"[UPLOAD_INVENTARIO] ✗ ERROR: {e}")
        import traceback
        traceback.print_exc()
        print(f"[UPLOAD_INVENTARIO] ===== FALLO =====\n")
        # Si falla, devolvemos el error
        return {
            "mensaje": f"❌ Error: {str(e)}",
            "archivo": {"uploader": session_id, "filename": save_name, "timestamp": timestamp},
            "parsed": []
        }
    
    msg = "✅ Inventario cargado correctamente" if parsed_rows else "⚠️ Archivo procesado pero sin datos"
    return {
        "mensaje": msg,
        "archivo": {"uploader": session_id, "filename": save_name, "timestamp": timestamp},
        "parsed": parsed_rows
    }


@app.post("/admin/usar_archivo")
async def admin_usar_archivo(request: Request, usuario: str, clear: bool = True) -> Dict[str, Any]:
    """Carga automáticamente la lista del usuario desde el archivo Inventario.xlsx.
    Si clear=True, limpia la lista anterior antes de cargar."""
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    if usuario not in usuarios:
        raise HTTPException(status_code=404, detail=f"Usuario {usuario} no encontrado")
    
    if df.empty:
        raise HTTPException(status_code=400, detail="Archivo Inventario.xlsx no cargado o vacío")
    
    try:
        # Limpiar lista si se solicita
        if clear and usuario in listas_por_usuario:
            listas_por_usuario[usuario] = []
        elif usuario not in listas_por_usuario:
            listas_por_usuario[usuario] = []
        
        # Obtener datos del usuario
        user_data = usuarios.get(usuario, {})
        legajo = user_data.get("legajo", "")
        
        # Procesar cada fila en df (Inventario.xlsx)
        for _, row in df.iterrows():
            codigo = str(row.get("codigo", "")).strip()
            descripcion = str(row.get("descripcion", "")).strip()
            stock = row.get("stock", "")
            ean_val = str(row.get("ean", "")).strip() if "ean" in df.columns else ""
            
            if not codigo:
                continue
            
            # Evitar duplicados
            if any(p.get("Codigo") == codigo for p in listas_por_usuario[usuario]):
                continue
            
            # Crear producto con fecha "hoy"
            hoy = datetime.today().strftime("%Y-%m-%d")
            estado_texto, dias_restantes = estado_vencimiento(hoy)
            
            listas_por_usuario[usuario].append({
                "Codigo": codigo,
                "Descripcion": descripcion,
                "EAN": ean_val,
                "Stock": stock,
                "FechaVencimiento": hoy,
                "Estado": estado_texto,
                "DiasRestantes": dias_restantes,
                "Usuario": usuario,
                "Legajo": legajo
            })
        
        guardar_listas()
        cantidad = len(listas_por_usuario[usuario])
        return {"mensaje": f"✅ Lista de {usuario} cargada con {cantidad} productos desde Inventario.xlsx", "lista": listas_por_usuario[usuario]}
    except Exception as e:
        print(f"Error cargando lista desde Inventario: {e}")
        raise HTTPException(status_code=500, detail=f"Error cargando lista: {str(e)}")


# --- EAN base ---
@app.post("/admin/upload_ean")
async def admin_upload_ean(request: Request, file: UploadFile = File(...)) -> Dict[str, Any]:
    """Carga un archivo que relaciona EAN con códigos y actualiza EAN.xlsx."""
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")

    user_dir = os.path.join(UPLOAD_DIR, session_id)
    os.makedirs(user_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    safe_name = (file.filename or "").replace('..', '')
    save_name = f"{timestamp}_{safe_name}"
    path = os.path.join(user_dir, save_name)
    contents = await file.read()
    with open(path, "wb") as f:
        f.write(contents)

    print(f"\n[UPLOAD_EAN] ===== Inicio =====")
    print(f"[UPLOAD_EAN] Archivo guardado en: {path}")

    global ean_df
    parsed = []
    
    try:
        # Leer archivo
        print(f"[UPLOAD_EAN] Leyendo archivo...")
        temp = pd.read_excel(path)
        print(f"[UPLOAD_EAN] Columnas originales: {list(temp.columns)}")
        
        # Normalizar columnas
        temp.columns = _normalize_columns(temp.columns)
        print(f"[UPLOAD_EAN] Columnas normalizadas: {list(temp.columns)}")
        
        # Validar que existan EAN y Código
        if 'ean' not in temp.columns or 'codigo' not in temp.columns:
            cols_faltantes = []
            if 'ean' not in temp.columns:
                cols_faltantes.append('EAN')
            if 'codigo' not in temp.columns:
                cols_faltantes.append('Código')
            raise ValueError(f"Faltan columnas: {', '.join(cols_faltantes)}")
        
        print(f"[UPLOAD_EAN] ✓ Columnas EAN y Código encontradas")
        
        # Limpiar datos (quitar '.0' resultado de números)
        temp['ean'] = temp['ean'].astype(str).apply(_strip_trailing_dot_zero)
        temp['codigo'] = temp['codigo'].astype(str).apply(_strip_trailing_dot_zero)
        # Enviar todos los registros para que la vista de EAN muestre la lista completa
        parsed = ean_df.to_dict(orient='records')
        print(f"[UPLOAD_EAN] ✓ ean_df actualizado: {len(parsed)} filas mostradas")
        print(f"[UPLOAD_EAN] ===== ✓ ÉXITO =====\n")
        
    except Exception as e:
        print(f"[UPLOAD_EAN] ✗ ERROR: {e}")
        import traceback
        traceback.print_exc()
        print(f"[UPLOAD_EAN] ===== FALLO =====\n")
        # Si falla, devolvemos el error
        return {
            "mensaje": f"❌ Error: {str(e)}",
            "archivo": {"uploader": session_id, "filename": save_name, "timestamp": timestamp},
            "parsed": []
        }
    
    msg = "✅ EAN cargado correctamente" if parsed else "⚠️ Archivo procesado pero sin datos"
    return {
        "mensaje": msg,
        "archivo": {"uploader": session_id, "filename": save_name, "timestamp": timestamp},
        "parsed": parsed
    }


@app.get("/admin/obtener_datos_busqueda")
async def admin_obtener_datos_busqueda(request: Request) -> Dict[str, List[Dict[str, Any]]]:
    """Devuelve TODOS los datos para autocompletado: Inventario.xlsx + EAN.xlsx.
    Combina datos de ambos archivos para búsqueda en cliente."""
    session_id = request.cookies.get("session_id")
    if not session_id:
        return {"items": []}
    try:
        items = []
        
        # Agregar datos del inventario base (Inventario.xlsx)
        if not df.empty:
            inv_items = df.to_dict(orient='records')
            print(f"[OBTENER_DATOS] Inventario tiene {len(inv_items)} items")
            items.extend(inv_items)
        
        # Agregar datos EAN (EAN.xlsx) - estos sobreescriben si hay duplicados por código
        if not ean_df.empty:
            ean_items = ean_df.to_dict(orient='records')
            print(f"[OBTENER_DATOS] EAN tiene {len(ean_items)} items")
            # Crear un dict de códigos EAN para fácil actualización
            ean_by_codigo = {str(r.get("codigo","")).strip(): r for r in ean_items}
            # Actualizar items que tienen EAN
            for item in items:
                codigo = str(item.get("codigo","")).strip()
                if codigo in ean_by_codigo:
                    item["ean"] = ean_by_codigo[codigo].get("ean", "")
            # Agregar items EAN que no están en inventario
            inv_codigos = {str(i.get("codigo","")).strip() for i in items}
            for ean_item in ean_items:
                codigo = str(ean_item.get("codigo","")).strip()
                if codigo not in inv_codigos:
                    items.append(ean_item)
        
        print(f"[OBTENER_DATOS] ✓ Total {len(items)} items para búsqueda")
        return {"items": items}
    except Exception as e:
        print(f"[OBTENER_DATOS] ✗ Error: {e}")
        import traceback
        traceback.print_exc()
        return {"items": []}


@app.get("/admin/obtener_ean")
async def admin_obtener_ean(request: Request) -> Dict[str, List[Dict[str, Any]]]:
    """Devuelve TODOS los datos para autocompletado: Inventario.xlsx + EAN.xlsx.
    Combina datos de ambos archivos para búsqueda en cliente."""
    session_id = request.cookies.get("session_id")
    if not session_id:
        print(f"[OBTENER_EAN] No hay session_id")
        return {"ean_items": []}
    try:
        ean_items = []
        
        # Agregar datos del inventario base (Inventario.xlsx)
        if not df.empty:
            inv_items = df.to_dict(orient='records')
            print(f"[OBTENER_EAN] Inventario tiene {len(inv_items)} items")
            ean_items.extend(inv_items)
        
        # Agregar/actualizar datos EAN (EAN.xlsx)
        if not ean_df.empty:
            ean_only = ean_df.to_dict(orient='records')
            print(f"[OBTENER_EAN] EAN tiene {len(ean_only)} items")
            # Crear un dict de códigos para actualizar EAN en inventario
            ean_by_codigo = {str(r.get("codigo","")).strip().upper(): r for r in ean_only}
            inv_codigos = {str(i.get("codigo","")).strip().upper() for i in ean_items}
            
            # Actualizar items que tienen EAN
            for item in ean_items:
                codigo = str(item.get("codigo","")).strip().upper()
                if codigo in ean_by_codigo:
                    item["ean"] = ean_by_codigo[codigo].get("ean", item.get("ean", ""))
            
            # Agregar items EAN que no están en inventario
            for ean_item in ean_only:
                codigo = str(ean_item.get("codigo","")).strip().upper()
                if codigo not in inv_codigos:
                    ean_items.append(ean_item)
        
        print(f"[OBTENER_EAN] ✓ Enviando {len(ean_items)} items totales")
        return {"ean_items": ean_items}
    except Exception as e:
        print(f"[OBTENER_EAN] ✗ Error: {e}")
        import traceback
        traceback.print_exc()
        return {"ean_items": []}


@app.get("/admin/obtener_inventario")
async def admin_obtener_inventario(request: Request) -> Dict[str, List[Dict[str, Any]]]:
    """Devuelve todos los datos del Inventario para gestión en el Admin."""
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        print(f"[OBTENER_INVENTARIO] No autorizado o sin session")
        return {"inventario_items": []}
    try:
        if df.empty:
            print(f"[OBTENER_INVENTARIO] df vacío")
            return {"inventario_items": []}
        
        inv_items = df.to_dict(orient='records')
        print(f"[OBTENER_INVENTARIO] ✓ Enviando {len(inv_items)} items del Inventario")
        return {"inventario_items": inv_items}
    except Exception as e:
        print(f"[OBTENER_INVENTARIO] ✗ Error: {e}")
        import traceback
        traceback.print_exc()
        return {"inventario_items": []}


@app.get("/debug_ean")
async def debug_ean() -> Dict[str, Any]:
    """Debug endpoint para ver estado de ean_df"""
    return {
        "ean_df_empty": ean_df.empty,
        "ean_df_shape": ean_df.shape if not ean_df.empty else None,
        "ean_df_columns": list(ean_df.columns) if not ean_df.empty else [],
        "ean_df_rows": ean_df.head(5).to_dict(orient='records') if not ean_df.empty else [],
        "ean_xlsx_exists": os.path.exists("EAN.xlsx")
    }


# ========== NUEVOS ENDPOINTS: ACTUALIZAR INVENTARIO Y EAN ==========

@app.post("/admin/actualizar_inventario")
async def admin_actualizar_inventario(request: Request, accion: str = "", codigo: str = "", codigo_nuevo: str = "", descripcion: str = "", stock: str = "") -> Dict[str, Any]:
    """Actualiza Inventario.xlsx directamente (agregar/editar/eliminar producto).
    - accion: "agregar" | "editar" | "eliminar"
    - codigo: código actual (clave)
    - codigo_nuevo: nuevo código (solo para editar)
    - descripcion: descripción del producto
    - stock: cantidad en stock (opcional)
    """
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    global df
    
    try:
        codigo = _strip_trailing_dot_zero(codigo.strip())
        codigo_nuevo = _strip_trailing_dot_zero(codigo_nuevo.strip())
        descripcion = descripcion.strip()
        
        if not codigo:
            return {"error": "El código es requerido", "data": []}
        
        if accion == "agregar":
            if not descripcion:
                return {"error": "La descripción es requerida al agregar", "data": []}
            
            # Verificar que no exista
            if not df.empty and (df['codigo'].astype(str) == codigo).any():
                return {"error": f"El código {codigo} ya existe", "data": df.to_dict(orient='records')}
            
            # Agregar fila
            nueva_fila = {
                'codigo': codigo,
                'descripcion': descripcion,
                'stock': stock if stock else 0
            }
            df = pd.concat([df, pd.DataFrame([nueva_fila])], ignore_index=True)
            df.to_excel('Inventario.xlsx', index=False)
            return {"mensaje": f"✅ Producto {codigo} agregado", "data": df.to_dict(orient='records')}
        
        elif accion == "editar":
            if df.empty:
                return {"error": "Inventario vacío", "data": []}
            
            idx = df[df['codigo'].astype(str) == codigo].index
            if idx.empty:
                return {"error": f"Código {codigo} no encontrado", "data": df.to_dict(orient='records')}
            
            idx = idx[0]

            # Permitir cambiar el código (clave) si se pasó uno nuevo
            if codigo_nuevo and codigo_nuevo != codigo:
                if not df.empty and (df['codigo'].astype(str) == codigo_nuevo).any():
                    return {"error": f"El código {codigo_nuevo} ya existe", "data": df.to_dict(orient='records')}
                df.at[idx, 'codigo'] = codigo_nuevo

            if descripcion:
                df.at[idx, 'descripcion'] = descripcion
            if stock:
                df.at[idx, 'stock'] = stock
            
            df.to_excel('Inventario.xlsx', index=False)
            return {"mensaje": f"✅ Producto {codigo} actualizado", "data": df.to_dict(orient='records')}
        
        elif accion == "eliminar":
            if df.empty:
                return {"error": "Inventario vacío", "data": []}
            
            idx = df[df['codigo'].astype(str) == codigo].index
            if idx.empty:
                return {"error": f"Código {codigo} no encontrado", "data": df.to_dict(orient='records')}
            
            df = df.drop(idx[0]).reset_index(drop=True)
            df.to_excel('Inventario.xlsx', index=False)
            return {"mensaje": f"✅ Producto {codigo} eliminado", "data": df.to_dict(orient='records')}
        
        else:
            return {"error": "Acción no válida (agregar/editar/eliminar)", "data": df.to_dict(orient='records')}
    
    except Exception as e:
        print(f"[ERROR actualizar_inventario]: {e}")
        return {"error": str(e), "data": df.to_dict(orient='records') if not df.empty else []}


@app.post("/admin/actualizar_ean")
async def admin_actualizar_ean(request: Request, accion: str = "", ean: str = "", nuevo_ean: str = "", codigo: str = "", descripcion: str = "") -> Dict[str, Any]:
    """Actualiza EAN.xlsx directamente (agregar/editar/eliminar mapeo EAN→Código).
    - accion: "agregar" | "editar" | "eliminar"
    - ean: EAN actual (clave)
    - nuevo_ean: nuevo EAN (solo para editar)
    - codigo: código asociado
    - descripcion: descripción del producto (opcional)
    """
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    global ean_df
    
    try:
        ean = _strip_trailing_dot_zero(ean.strip())
        nuevo_ean = _strip_trailing_dot_zero(nuevo_ean.strip())
        codigo = _strip_trailing_dot_zero(codigo.strip())
        descripcion = descripcion.strip()
        
        if not ean or not codigo:
            return {"error": "EAN y Código son requeridos", "data": []}
        
        if accion == "agregar":
            # Verificar que no exista
            if not ean_df.empty and (ean_df['ean'].astype(str) == ean).any():
                return {"error": f"El EAN {ean} ya existe", "data": ean_df.to_dict(orient='records')}
            
            # Agregar fila
            nueva_fila = {
                'ean': ean,
                'codigo': codigo,
                'descripcion': descripcion if descripcion else ''
            }
            ean_df = pd.concat([ean_df, pd.DataFrame([nueva_fila])], ignore_index=True)
            ean_df.to_excel('EAN.xlsx', index=False)
            return {"mensaje": f"✅ EAN {ean} agregado", "data": ean_df.to_dict(orient='records')}
        
        elif accion == "editar":
            if ean_df.empty:
                return {"error": "Base EAN vacía", "data": []}
            
            idx = ean_df[ean_df['ean'].astype(str) == ean].index
            if idx.empty:
                return {"error": f"EAN {ean} no encontrado", "data": ean_df.to_dict(orient='records')}
            
            idx = idx[0]

            # Permitir cambiar el EAN (clave)
            if nuevo_ean and nuevo_ean != ean:
                if not ean_df.empty and (ean_df['ean'].astype(str) == nuevo_ean).any():
                    return {"error": f"El EAN {nuevo_ean} ya existe", "data": ean_df.to_dict(orient='records')}
                ean_df.at[idx, 'ean'] = nuevo_ean

            if codigo:
                ean_df.at[idx, 'codigo'] = codigo
            if descripcion:
                ean_df.at[idx, 'descripcion'] = descripcion
            
            ean_df.to_excel('EAN.xlsx', index=False)
            return {"mensaje": f"✅ EAN {ean} actualizado", "data": ean_df.to_dict(orient='records')}
        
        elif accion == "eliminar":
            if ean_df.empty:
                return {"error": "Base EAN vacía", "data": []}
            
            idx = ean_df[ean_df['ean'].astype(str) == ean].index
            if idx.empty:
                return {"error": f"EAN {ean} no encontrado", "data": ean_df.to_dict(orient='records')}
            
            ean_df = ean_df.drop(idx[0]).reset_index(drop=True)
            ean_df.to_excel('EAN.xlsx', index=False)
            return {"mensaje": f"✅ EAN {ean} eliminado", "data": ean_df.to_dict(orient='records')}
        
        else:
            return {"error": "Acción no válida (agregar/editar/eliminar)", "data": ean_df.to_dict(orient='records')}
    
    except Exception as e:
        print(f"[ERROR actualizar_ean]: {e}")
        return {"error": str(e), "data": ean_df.to_dict(orient='records') if not ean_df.empty else []}


@app.get("/buscar_from_ean_e_inventario")
async def buscar_ean_inventario(q: str = "") -> Dict[str, Any]:
    """Busca en AMBOS: ean_df (EAN.xlsx) e inventario (Inventario.xlsx)."""
    if not q or len(q.strip()) < 1:
        return {"resultados": []}
    
    q_norm = _norm_str(q)
    # eliminar posible ".0" al final de la consulta
    if q_norm.endswith('.0'):
        q_norm = q_norm[:-2]
    resultados = []
    visto = set()
    
    print(f"[BUSCAR] Query normalizado: '{q_norm}' (original: '{q}')")
    
    # Buscar en EAN.xlsx
    if not ean_df.empty:
        print(f"[BUSCAR] Buscando en ean_df ({len(ean_df)} filas)...")
        for idx, row in ean_df.iterrows():
            ean_val = _norm_str(str(row.get('ean', '') or ''))
            cod_val = _norm_str(str(row.get('codigo', '') or ''))
            desc_val = _norm_str(str(row.get('descripcion', '') or ''))
            
            if q_norm in ean_val or q_norm in cod_val or q_norm in desc_val:
                clave = (ean_val, cod_val)
                if clave not in visto:
                    visto.add(clave)
                    resultado = {
                        'ean': _strip_trailing_dot_zero(str(row.get('ean', '') or '')),
                        'codigo': _strip_trailing_dot_zero(str(row.get('codigo', '') or '')),
                        'descripcion': str(row.get('descripcion', '') or '')
                    }
                    resultados.append(resultado)
                    print(f"[BUSCAR] ✓ Encontrado en EAN: {resultado}")
    else:
        print(f"[BUSCAR] ean_df está vacío")
    
    # Buscar en Inventario.xlsx
    if not df.empty:
        print(f"[BUSCAR] Buscando en df inventario ({len(df)} filas)...")
        for idx, row in df.iterrows():
            cod_val = _norm_str(str(row.get('codigo', '') or ''))
            desc_val = _norm_str(str(row.get('descripcion', '') or ''))
            
            if q_norm in cod_val or q_norm in desc_val:
                clave = (q_norm, cod_val)
                if clave not in visto:
                    visto.add(clave)
                    resultado = {
                        'ean': '',
                        'codigo': str(row.get('codigo', '') or ''),
                        'descripcion': str(row.get('descripcion', '') or '')
                    }
                    resultados.append(resultado)
                    print(f"[BUSCAR] ✓ Encontrado en inventario: {resultado}")
    else:
        print(f"[BUSCAR] df inventario está vacío")
    
    print(f"[BUSCAR] Resultados totales: {len(resultados)}")
    return {"resultados": resultados[:20]}


@app.post("/admin/usar_archivo")
async def admin_usar_archivo(request: Request, user: str, filename: str, clear: bool = True) -> Dict[str, Any]:
    """Establece un archivo subido como Inventario activo.
    Parámetros: user, filename, clear (si True borra listas de usuario antiguas).
    """
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")

    path = os.path.join(UPLOAD_DIR, user, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    global df, listas_por_usuario
    parsed_rows: List[Dict[str, Any]] = []
    try:
        temp = pd.read_excel(path)
        # normalizar encabezados para quitar acentos/espacios y pasar a minúsculas
        temp.columns = _normalize_columns(temp.columns)
        if 'codigo' in temp.columns and 'descripcion' in temp.columns:
            temp.to_excel('Inventario.xlsx', index=False)
            df = temp
            parsed_rows = df.head(50).to_dict(orient='records')
            # también preparar la lista del administrador con los primeros elementos
            listas_por_usuario[session_id] = [
                {"Codigo": r.get("codigo", ""), "Descripcion": r.get("descripcion", ""),
                 "FechaVencimiento": "", "Stock": r.get("stock", ""),
                 "Usuario": session_id, "Legajo": usuarios.get(session_id, {}).get("legajo", "")}
                for r in parsed_rows
            ]
            guardar_listas()
            if clear:
                # save previous lists to a backup file so admin can restore if needed
                try:
                    with open('listas_backup.json', 'w', encoding='utf-8') as bk:
                        json.dump(listas_por_usuario, bk, ensure_ascii=False, indent=2)
                except Exception:
                    pass
                listas_por_usuario.clear()
            return {"mensaje": "Inventario actualizado", "parsed": parsed_rows}
        else:
            raise HTTPException(status_code=400, detail="El archivo no contiene las columnas mínimas")
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error al aplicar archivo como inventario: {e}")
        raise HTTPException(status_code=500, detail="Error procesando el archivo")


@app.get("/admin/listar_archivos")
async def admin_listar_archivos(request: Request) -> Dict[str, List[Dict[str, Any]]]:
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")

    archivos: List[Dict[str, Any]] = []
    for user in os.listdir(UPLOAD_DIR):
        user_dir = os.path.join(UPLOAD_DIR, user)
        if not os.path.isdir(user_dir):
            continue
        for fname in os.listdir(user_dir):
            fpath = os.path.join(user_dir, fname)
            try:
                mtime = os.path.getmtime(fpath)
            except OSError:
                continue
            archivos.append({
                "uploader": user,
                "filename": fname,
                "path": fpath,
                "mtime": mtime
            })

    archivos.sort(key=lambda x: x.get("mtime", 0))  # de más antiguo a más reciente

    result: List[Dict[str, Any]] = []
    for a in archivos:
        dt = datetime.fromtimestamp(a.get("mtime", 0))
        result.append({
            "uploader": a["uploader"],
            "filename": a["filename"],
            "fecha": dt.strftime("%d-%m-%Y"),
            "hora": dt.strftime("%H:%M:%S"),
            "size": os.path.getsize(a["path"]) if os.path.exists(a["path"]) else 0
        })

    return {"archivos": result}


@app.get("/admin/descargar_lista_unificada")
async def admin_descargar_lista_unificada(request: Request):
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    # Recopilar todas las listas de todos los usuarios
    datos_excel = []
    
    for user_id, lista in listas_por_usuario.items():
        for p in lista:
            datos_excel.append({
                "Usuario": p.get("Usuario", ""),
                "Legajo": p.get("Legajo", ""),
                "Código": p.get("Codigo", ""),
                "Descripción": p.get("Descripcion", ""),
                "Fecha Vencimiento": p.get("FechaVencimiento", ""),
                "Stock": p.get("Stock", "-")
            })
    
    if not datos_excel:
        raise HTTPException(status_code=400, detail="No hay datos para descargar")
    
    # Crear Excel con estilo y fórmulas usando openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos Unificados"

    # columnas: Usuario, Legajo, Código, Descripción, Fecha Vencimiento, Stock, Días Restantes, Estado
    headers = ["Usuario", "Legajo", "Código", "Descripción", "Fecha Vencimiento", "Stock", "Días Restantes", "Estado"]
    # Estilo de encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Escribir filas
    row = 2
    for p in datos_excel:
        ws.cell(row=row, column=1, value=p.get("Usuario", ""))
        ws.cell(row=row, column=2, value=p.get("Legajo", ""))
        ws.cell(row=row, column=3, value=p.get("Código", ""))
        ws.cell(row=row, column=4, value=p.get("Descripción", ""))
        fecha_str = p.get("Fecha Vencimiento", "")
        if fecha_str:
            try:
                fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
                fecha_cell = ws.cell(row=row, column=5, value=fecha_obj)
                fecha_cell.number_format = 'DD/MM/YYYY'
            except:
                ws.cell(row=row, column=5, value=fecha_str)
        else:
            ws.cell(row=row, column=5, value="")
        ws.cell(row=row, column=6, value=p.get("Stock", "-"))
        row += 1

    max_row = ws.max_row

    # Insertar fórmulas para Días Restantes (col G) y Estado (col H)
    for r in range(2, max_row + 1):
        date_cell = f"E{r}"
        days_cell = f"G{r}"
        status_cell = f"H{r}"
        # Días restantes = Fecha - HOY()
        ws[days_cell].value = f"=INT({date_cell}-TODAY())"
        # Estado según días
        ws[status_cell].value = (
            f"=IF({days_cell}<=0,\"Vencido\",IF({days_cell}<=7,\"Crítico (<7 días)\",\"Correcto\"))"
        )

    # Aplicar formato condicional
    red_fill = PatternFill(start_color="F87171", end_color="F87171", fill_type="solid")
    yellow_fill = PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid")
    green_fill = PatternFill(start_color="34D399", end_color="34D399", fill_type="solid")

    if max_row >= 2:
        ws.conditional_formatting.add(f"G2:G{max_row}", CellIsRule(operator='lessThanOrEqual', formula=['0'], stopIfTrue=True, fill=red_fill))
        ws.conditional_formatting.add(f"G2:G{max_row}", CellIsRule(operator='between', formula=['1','7'], stopIfTrue=True, fill=yellow_fill))
        ws.conditional_formatting.add(f"G2:G{max_row}", CellIsRule(operator='greaterThan', formula=['7'], stopIfTrue=True, fill=green_fill))
        ws.conditional_formatting.add(f"H2:H{max_row}", FormulaRule(formula=[f"$G2<=0"], stopIfTrue=True, fill=red_fill))
        ws.conditional_formatting.add(f"H2:H{max_row}", FormulaRule(formula=[f"AND($G2>0,$G2<=7)"], stopIfTrue=True, fill=yellow_fill))
        ws.conditional_formatting.add(f"H2:H{max_row}", FormulaRule(formula=[f"$G2>7"], stopIfTrue=True, fill=green_fill))

    # Colorear columnas Usuario y Legajo
    user_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    leg_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    for r in range(2, max_row + 1):
        ws.cell(row=r, column=1).fill = user_fill
        ws.cell(row=r, column=2).fill = leg_fill

    # Ajustar anchos
    widths = [16, 12, 15, 40, 16, 10, 14, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Autofiltro
    if max_row >= 2:
        ws.auto_filter.ref = f"A1:H{max_row}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {"Content-Disposition": "attachment; filename=lista_unificada_usuarios.xlsx"}
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@app.get("/admin/descargar_archivo")
async def admin_descargar_archivo(user: str, filename: str, request: Request):
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")

    path = os.path.join(UPLOAD_DIR, user, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    def iterfile():
        with open(path, "rb") as f:
            while True:
                chunk = f.read(8192)
                if not chunk:
                    break
                yield chunk

    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(iterfile(), media_type="application/octet-stream", headers=headers)


@app.get("/admin/descargar_todos")
async def admin_descargar_todos(request: Request):
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")

    entries: List[Tuple[str, str, float]] = []
    for user in os.listdir(UPLOAD_DIR):
        user_dir = os.path.join(UPLOAD_DIR, user)
        if not os.path.isdir(user_dir):
            continue
        for fname in os.listdir(user_dir):
            fpath = os.path.join(user_dir, fname)
            try:
                mtime = os.path.getmtime(fpath)
            except OSError:
                continue
            entries.append((fpath, os.path.join(user, fname), mtime))

    entries.sort(key=lambda x: x[2])

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for fpath, arcname, _ in entries:
            try:
                z.write(fpath, arcname)
            except OSError:
                continue
    buf.seek(0)
    headers = {"Content-Disposition": "attachment; filename=todos_archivos.zip"}
    return StreamingResponse(buf, media_type="application/zip", headers=headers)

@app.get("/admin/obtener_lista_usuario")
async def admin_obtener_lista_usuario(user: str, request: Request) -> Dict[str, Any]:
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    lista = listas_por_usuario.get(user, [])
    return {"lista": lista}

@app.get("/admin/descargar_lista_usuario")
async def admin_descargar_lista_usuario(user: str, request: Request):
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    excel_path = f"excels/{user}.xlsx"
    if os.path.exists(excel_path):
        # Devolver el Excel guardado persistentemente
        def iterfile():
            with open(excel_path, "rb") as f:
                while True:
                    chunk = f.read(8192)
                    if not chunk:
                        break
                    yield chunk
        headers = {"Content-Disposition": f"attachment; filename=lista_{user}.xlsx"}
        return StreamingResponse(iterfile(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    else:
        # Si no existe, generar desde la lista en memoria
        lista = listas_por_usuario.get(user, [])
        if not lista:
            raise HTTPException(status_code=400, detail="No hay datos para ese usuario")
        
        # ordenar por DiasRestantes (asc)
        lista = sorted(lista, key=lambda p: p.get("DiasRestantes", float('inf')))
        
        # Crear Excel similar al guardar_lista, columnas reordenadas
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"
        headers = ["Código", "Descripción", "Fecha Vencimiento", "Estado", "Legajo", "Usuario", "Stock", "Días Restantes"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 2
        for p in lista:
            ws.cell(row=row, column=1, value=p.get("Codigo", ""))
            ws.cell(row=row, column=2, value=p.get("Descripcion", ""))
            if p.get("FechaVencimiento"):
                fecha_obj = datetime.strptime(p["FechaVencimiento"], "%Y-%m-%d")
                fecha_cell = ws.cell(row=row, column=3, value=fecha_obj)
                fecha_cell.number_format = 'DD/MM/YYYY'
            else:
                ws.cell(row=row, column=3, value="")
            ws.cell(row=row, column=4, value=p.get("Estado", ""))
            ws.cell(row=row, column=5, value=p.get("Legajo", ""))
            ws.cell(row=row, column=6, value=p.get("Usuario", ""))
            ws.cell(row=row, column=7, value=p.get("Stock", "-"))
            ws.cell(row=row, column=8, value=p.get("DiasRestantes", ""))
            row += 1
        
        max_row = ws.max_row
        red_fill = PatternFill(start_color="F87171", end_color="F87171", fill_type="solid")
        yellow_fill = PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid")
        green_fill = PatternFill(start_color="34D399", end_color="34D399", fill_type="solid")
        if max_row >= 2:
            ws.conditional_formatting.add(f"H2:H{max_row}", CellIsRule(operator='lessThanOrEqual', formula=['0'], stopIfTrue=True, fill=red_fill))
            ws.conditional_formatting.add(f"H2:H{max_row}", CellIsRule(operator='between', formula=['1','7'], stopIfTrue=True, fill=yellow_fill))
            ws.conditional_formatting.add(f"H2:H{max_row}", CellIsRule(operator='greaterThan', formula=['7'], stopIfTrue=True, fill=green_fill))
            ws.conditional_formatting.add(f"D2:D{max_row}", FormulaRule(formula=["$H2<=0"], stopIfTrue=True, fill=red_fill))
            ws.conditional_formatting.add(f"D2:D{max_row}", FormulaRule(formula=["AND($H2>0,$H2<=7)"], stopIfTrue=True, fill=yellow_fill))
            ws.conditional_formatting.add(f"D2:D{max_row}", FormulaRule(formula=["$H2>7"], stopIfTrue=True, fill=green_fill))
        
        user_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        leg_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=6).fill = user_fill
            ws.cell(row=r, column=5).fill = leg_fill
        
        widths = [40, 15, 16, 10, 14, 20, 16, 12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        
        if max_row >= 2:
            ws.auto_filter.ref = f"A1:H{max_row}"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        headers = {"Content-Disposition": f"attachment; filename=lista_{user}.xlsx"}
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@app.post("/admin/borrar_archivo")
async def admin_borrar_archivo(user: str, filename: str, request: Request) -> Dict[str, Any]:
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")

    path = os.path.join(UPLOAD_DIR, user, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    try:
        os.remove(path)
        # Si el directorio del usuario quedó vacío, intentar eliminarlo
        user_dir = os.path.join(UPLOAD_DIR, user)
        if os.path.isdir(user_dir) and not os.listdir(user_dir):
            try:
                os.rmdir(user_dir)
            except Exception:
                pass
        return {"mensaje": "Archivo eliminado"}
    except Exception as e:
        print(f"Error al eliminar archivo: {e}")
        raise HTTPException(status_code=500, detail="Error eliminando archivo")


@app.post("/borrar_archivo")
async def borrar_archivo(filename: str, request: Request) -> Dict[str, Any]:
    """Permite a usuarios borrar sus propios archivos subidos"""
    session_id = request.cookies.get("session_id")
    if not session_id:
        raise HTTPException(status_code=403, detail="No autenticado")

    # Solo puede borrar archivos de su propio directorio
    path = os.path.join(UPLOAD_DIR, session_id, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    try:
        os.remove(path)
        # Si el directorio del usuario quedó vacío, intentar eliminarlo
        user_dir = os.path.join(UPLOAD_DIR, session_id)
        if os.path.isdir(user_dir) and not os.listdir(user_dir):
            try:
                os.rmdir(user_dir)
            except Exception:
                pass
        return {"mensaje": "Archivo eliminado"}
    except Exception as e:
        print(f"Error al eliminar archivo: {e}")
        raise HTTPException(status_code=500, detail="Error eliminando archivo")


@app.get("/listar_mis_archivos")
async def listar_mis_archivos(request: Request) -> Dict[str, List[Dict[str, Any]]]:
    """Permite a usuarios listar sus propios archivos subidos"""
    session_id = request.cookies.get("session_id")
    if not session_id:
        raise HTTPException(status_code=403, detail="No autenticado")

    user_dir = os.path.join(UPLOAD_DIR, session_id)
    archivos: List[Dict[str, Any]] = []

    if os.path.exists(user_dir):
        for fname in os.listdir(user_dir):
            fpath = os.path.join(user_dir, fname)
            try:
                mtime = os.path.getmtime(fpath)
                size = os.path.getsize(fpath)
            except OSError:
                continue
            dt = datetime.fromtimestamp(mtime)
            archivos.append({
                "filename": fname,
                "fecha": dt.strftime("%d-%m-%Y"),
                "hora": dt.strftime("%H:%M:%S"),
                "size": size
            })

    archivos.sort(key=lambda x: x.get("mtime", 0), reverse=True)  # más reciente primero
    return {"archivos": archivos}


@app.post("/session/close")
async def session_close(request: Request) -> Dict[str, Any]:
    """Cierra la sesión del usuario actual. JSON: {"clear": true|false}
    Si clear es True se eliminan las listas del usuario antes de cerrar.
    """
    session_id = request.cookies.get("session_id")
    if not session_id:
        raise HTTPException(status_code=403, detail="No autenticado")

    data = await request.json()
    clear = bool(data.get("clear", False))

    try:
        if clear:
            # eliminar lista del usuario
            listas_por_usuario.pop(session_id, None)
            guardar_listas()
        # enviar correo de cierre de sesión
        try:
            send_email_to_user(session_id, "Sesión cerrada", f"Usuario {session_id} ha cerrado sesión. Clear={clear}.")
        except Exception:
            pass
        # devolver indicación; el frontend debe llamar a /logout para eliminar cookie
        return {"mensaje": "Sesión preparada para cerrar", "clear": clear}
    except Exception as e:
        print(f"Error cerrando sesión: {e}")
        raise HTTPException(status_code=500, detail="Error procesando cierre de sesión")


def _agregar_ean_a_resultado(producto_dict: Dict[str, Any]) -> Dict[str, Any]:
    """Agrega el EAN a un producto desde ean_df si es que existe."""
    try:
        codigo = str(producto_dict.get("codigo", "")).strip()
        if codigo and not ean_df.empty:
            ean_match = ean_df[ean_df["codigo"].astype(str).str.strip().str.upper() == codigo.upper()]
            if not ean_match.empty:
                producto_dict["ean"] = str(ean_match.iloc[0]["ean"])
            else:
                producto_dict["ean"] = ""
        else:
            producto_dict["ean"] = ""
    except Exception as e:
        print(f"Error añadiendo EAN: {e}")
        producto_dict["ean"] = ""
    return producto_dict

@app.get("/buscar_producto")
async def buscar_producto(code: Optional[str] = None, codigo: Optional[str] = None,
                          descripcion: Optional[str] = None, articulo: Optional[str] = None,
                          nombre: Optional[str] = None, ean: Optional[str] = None,
                          fecha: Optional[str] = None, limit: int = 50) -> Dict[str, List[Any]]:
    """Buscar en el Inventario.
    - `code`/`codigo`: coincidencia exacta de código.
    - `ean`: buscar por EAN y devolver el código asociado.
    - `descripcion`/`articulo`/`nombre`: búsqueda de texto en la descripción.
    Se pueden combinar filtros. Devuelve hasta `limit` coincidencias.
    """
    # alias de texto
    if articulo and not descripcion:
        descripcion = articulo
    if nombre and not descripcion:
        descripcion = nombre

    # alias de código
    if codigo and not code:
        code = codigo

    # normalizador local
    def norm(x):
        return _norm_str(x)
    
    # Función para comparar códigos ignorando ceros al inicio
    def codigo_match(codigo_a, codigo_b):
        """Compara dos códigos ignorando ceros a la izquierda"""
        a = _strip_leading_zeros(codigo_a)
        b = _strip_leading_zeros(codigo_b)
        return norm(a).upper() == norm(b).upper()

    # 1. Búsqueda por EAN (convertir a código)
    if ean:
        clave = norm(ean).upper()
        try:
            if not ean_df.empty:
                # Buscar EAN ignorando ceros
                for idx, row in ean_df.iterrows():
                    if norm(_strip_leading_zeros(row.get("ean", ""))).upper() == clave:
                        mapped_code = str(row.get("codigo", ""))
                        # Buscar el código en Inventario, ignorando ceros
                        for pidx, prow in df.iterrows():
                            if codigo_match(prow.get("codigo", ""), mapped_code):
                                res = prow.to_dict()
                                res = _agregar_ean_a_resultado(res)
                                return {"matches": [res]}
        except Exception as e:
            print(f"Error buscando por EAN: {e}")

    # 2. Búsqueda por código
    if code:
        try:
            matches = []
            for idx, row in df.iterrows():
                if codigo_match(row.get("codigo", ""), code):
                    res = row.to_dict()
                    res = _agregar_ean_a_resultado(res)
                    matches.append(res)
            if matches:
                return {"matches": matches[:limit]}
            # Si no encuentra en inventario, buscar en EAN
            if not matches and not ean_df.empty:
                for idx, row in ean_df.iterrows():
                    if codigo_match(row.get("ean", ""), code):
                        mapped = str(row.get("codigo", ""))
                        for pidx, prow in df.iterrows():
                            if codigo_match(prow.get("codigo", ""), mapped):
                                res = prow.to_dict()
                                res = _agregar_ean_a_resultado(res)
                                matches.append(res)
                        break
            return {"matches": matches[:limit] if matches else []}
        except Exception as e:
            print(f"Error buscando por código: {e}")
            return {"matches": []}

    # 3. Búsqueda por descripción solamente
    if descripcion:
        term = norm(descripcion)
        try:
            mask = df["descripcion"].astype(str).apply(norm).str.contains(term, case=False, na=False)
            # si llega parámetro fecha y existe columna, filtrar
            if fecha and fecha in df.columns:
                mask = mask & (df[fecha] == fecha)
            res = df[mask].head(limit).to_dict(orient="records")
            res = [_agregar_ean_a_resultado(r) for r in res]
            return {"matches": res}
        except Exception as e:
            print(f"Error buscando por descripción: {e}")
            return {"matches": []}

    return {"matches": []}

# --- Productos ---
@app.post("/agregar_producto")
async def agregar_producto(request: Request) -> Dict[str, Any]:
    data = await request.json()
    session_id = request.cookies.get("session_id")
    if not session_id:
        raise HTTPException(status_code=403, detail="No autenticado")
    if session_id not in listas_por_usuario:
        listas_por_usuario[session_id] = []

    codigo = data.get("codigo")
    descripcion = data.get("descripcion", "")
    ean = data.get("ean", "")
    fecha_vencimiento = data.get("fecha_vencimiento")

    if not codigo or not fecha_vencimiento:
        raise HTTPException(status_code=400, detail="Debe ingresar código y fecha")

    # Convert codigo to string in case it comes as integer from JSON
    codigo_str = str(codigo).strip().upper()
    
    producto = df[df["codigo"].astype(str).str.strip().str.upper() == codigo_str]
    stock = ""
    if not producto.empty:
        datos = producto.to_dict(orient="records")[0]
        stock = datos.get("stock", "")
        if not descripcion:
            descripcion = datos.get("descripcion", "")

    # Si no se proporcionó EAN, buscarlo en ean_df
    if not ean and not ean_df.empty:
        ean_match = ean_df[ean_df["codigo"].astype(str).str.strip().str.upper() == codigo_str]
        if not ean_match.empty:
            ean = str(ean_match.iloc[0]["ean"])

    # Evitar introducir el mismo código dos veces (formatos distintos, ceros, mayúsculas)
    for p in listas_por_usuario[session_id]:
        if str(p.get("Codigo", "")).strip().upper() == codigo_str:
            raise HTTPException(status_code=400, detail="Producto ya agregado")

    estado_texto, dias_restantes = estado_vencimiento(fecha_vencimiento)
    if estado_texto == "Fecha inválida":
        raise HTTPException(status_code=400, detail="Fecha de vencimiento inválida")

    legajo = usuarios[session_id].get("legajo", "")

    # Insertar al inicio para que el último agregado quede primero en la lista
    listas_por_usuario[session_id].insert(0, {
        "Codigo": codigo,
        "Descripcion": descripcion,
        "EAN": ean,
        "Stock": stock,
        "FechaVencimiento": fecha_vencimiento,
        "Estado": estado_texto,
        "DiasRestantes": dias_restantes,
        "Usuario": session_id,
        "Legajo": legajo
    })

    guardar_listas()
    return {"mensaje": "Producto agregado", "lista": listas_por_usuario[session_id]}

@app.post("/admin/cargar_lista_ean")
async def admin_cargar_lista_ean(request: Request, usuario: str, clear: bool = True) -> Dict[str, Any]:
    """Carga automáticamente una lista del usuario desde el archivo EAN.xlsx.
    Si clear=True, limpia la lista anterior antes de cargar.
    """
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    if usuario not in usuarios:
        raise HTTPException(status_code=404, detail=f"Usuario {usuario} no encontrado")
    
    if ean_df.empty:
        raise HTTPException(status_code=400, detail="Archivo EAN.xlsx no cargado o vacío")
    
    try:
        # Limpiar lista si se solicita
        if clear and usuario in listas_por_usuario:
            listas_por_usuario[usuario] = []
        elif usuario not in listas_por_usuario:
            listas_por_usuario[usuario] = []
        
        # Obtener datos del usuario
        user_data = usuarios.get(usuario, {})
        legajo = user_data.get("legajo", "")
        
        # Procesar cada fila en ean_df
        for _, row in ean_df.iterrows():
            codigo = str(row.get("codigo", "")).strip()
            ean_val = str(row.get("ean", "")).strip()
            
            if not codigo:
                continue
            
            # Evitar duplicados
            if any(p.get("Codigo") == codigo for p in listas_por_usuario[usuario]):
                continue
            
            # Buscar producto en df
            producto = df[df["codigo"].astype(str).str.strip().str.upper() == codigo.upper()]
            stock = ""
            descripcion = ""
            if not producto.empty:
                datos = producto.to_dict(orient="records")[0]
                stock = datos.get("stock", "")
                descripcion = datos.get("descripcion", "")
            
            # Crear producto con fecha "hoy"
            hoy = datetime.today().strftime("%Y-%m-%d")
            estado_texto, dias_restantes = estado_vencimiento(hoy)
            
            listas_por_usuario[usuario].append({
                "Codigo": codigo,
                "Descripcion": descripcion,
                "EAN": ean_val,
                "Stock": stock,
                "FechaVencimiento": hoy,
                "Estado": estado_texto,
                "DiasRestantes": dias_restantes,
                "Usuario": usuario,
                "Legajo": legajo
            })
        
        guardar_listas()
        cantidad = len(listas_por_usuario[usuario])
        return {"mensaje": f"Lista de {usuario} cargada con {cantidad} productos desde EAN.xlsx", "lista": listas_por_usuario[usuario]}
    except Exception as e:
        print(f"Error cargando lista desde EAN: {e}")
        raise HTTPException(status_code=500, detail=f"Error cargando lista: {str(e)}")

@app.post("/admin/restaurar_listas")
async def admin_restaurar_listas(request: Request) -> Dict[str, Any]:
    """Carga listas de usuario desde el último respaldo (listas_backup.json).
    Sólo administradores pueden ejecutar esta acción."""
    session_id = request.cookies.get("session_id")
    if not session_id or usuarios.get(session_id, {}).get("rol") != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    if os.path.exists('listas_backup.json'):
        try:
            with open('listas_backup.json', 'r', encoding='utf-8') as bk:
                data = json.load(bk)
                # asegurarnos de que sea un dict
                if isinstance(data, dict):
                    listas_por_usuario.clear()
                    listas_por_usuario.update(data)
                    guardar_listas()
                    return {"mensaje": "Listas restauradas desde respaldo"}
        except Exception as e:
            print(f"Error restaurando listas: {e}")
    raise HTTPException(status_code=404, detail="No hay respaldo disponible")

@app.delete("/borrar_lista_completa")
async def borrar_lista_completa(request: Request) -> Dict[str, Any]:
    session_id = request.cookies.get("session_id")
    if not session_id:
        raise HTTPException(status_code=403, detail="No autenticado")
    listas_por_usuario[session_id] = []
    guardar_listas()
    return {"mensaje": "Lista completa eliminada", "lista": []}

@app.post("/guardar_lista")
async def guardar_lista(request: Request):
    print("/guardar_lista called, cookies=", request.cookies)
    
    session_id = request.cookies.get("session_id")
    if not session_id:
        raise HTTPException(status_code=403, detail="No autenticado")

    lista = listas_por_usuario.get(session_id, [])
    if not lista:
        raise HTTPException(status_code=400, detail="La lista está vacía")

    # ordenar por DiasRestantes ascendente (si está presente)
    lista = sorted(lista, key=lambda p: p.get("DiasRestantes", float('inf')))

    # Obtener usuario para guardar el Excel
    sesion = sesiones.get(session_id)
    if not sesion:
        # Si el servidor se reinició pero el navegador aún mantiene la cookie,
        # permitimos seguir guardando la lista usando el id de sesión como usuario.
        sesion = {"usuario": session_id}
    usuario = sesion.get("usuario", "desconocido")

    # Crear Excel con estilo y fórmulas usando openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # columnas: Código, Descripción, Fecha Vencimiento, Estado, Legajo, Usuario, Stock, Días Restantes
    headers = ["Código", "Descripción", "Fecha Vencimiento", "Estado", "Legajo", "Usuario", "Stock", "Días Restantes"]
    # Estilo de encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Escribir filas y dejar espacio para fórmulas
    row = 2
    for p in lista:
        fecha_obj = datetime.strptime(p["FechaVencimiento"], "%Y-%m-%d") if p.get("FechaVencimiento") else None
        ws.cell(row=row, column=1, value=p.get("Codigo", ""))
        ws.cell(row=row, column=2, value=p.get("Descripcion", ""))
        if fecha_obj:
            fecha_cell = ws.cell(row=row, column=3, value=fecha_obj)
            fecha_cell.number_format = 'DD/MM/YYYY'
        else:
            ws.cell(row=row, column=3, value="")
        # Estado en columna 4
        ws.cell(row=row, column=4, value=p.get("Estado", ""))
        # Legajo y Usuario
        ws.cell(row=row, column=5, value=p.get("Legajo", ""))
        ws.cell(row=row, column=6, value=p.get("Usuario", ""))
        # Stock y DiasRestantes al final
        stock_val = p.get("Stock", "")
        stock_cell = ws.cell(row=row, column=7, value=stock_val)
        # Si el stock es numérico, mostrar con 3 decimales (por ejemplo, kilos)
        try:
            stock_float = float(stock_val)
            stock_cell.value = stock_float
            stock_cell.number_format = '0.000'
        except Exception:
            # dejar tal cual (por ejemplo, vacío o texto)
            pass
        ws.cell(row=row, column=8, value=p.get("DiasRestantes", 0))
        row += 1

    max_row = ws.max_row

    # Ya tenemos valores de Días Restantes en la lista; no hace falta fórmulas
    # Aplicar formato condicional para días (col H) y Estado (col D)
    red_fill = PatternFill(start_color="F87171", end_color="F87171", fill_type="solid")
    yellow_fill = PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid")
    green_fill = PatternFill(start_color="34D399", end_color="34D399", fill_type="solid")

    if max_row >= 2:
        # dias restantes está en columna 8 (H)
        ws.conditional_formatting.add(f"H2:H{max_row}", CellIsRule(operator='lessThanOrEqual', formula=['0'], stopIfTrue=True, fill=red_fill))
        ws.conditional_formatting.add(f"H2:H{max_row}", CellIsRule(operator='between', formula=['1','7'], stopIfTrue=True, fill=yellow_fill))
        ws.conditional_formatting.add(f"H2:H{max_row}", CellIsRule(operator='greaterThan', formula=['7'], stopIfTrue=True, fill=green_fill))
        # Estado en columna D coloreado según dias restantes
        ws.conditional_formatting.add(f"D2:D{max_row}", FormulaRule(formula=["$H2<=0"], stopIfTrue=True, fill=red_fill))
        ws.conditional_formatting.add(f"D2:D{max_row}", FormulaRule(formula=["AND($H2>0,$H2<=7)"], stopIfTrue=True, fill=yellow_fill))
        ws.conditional_formatting.add(f"D2:D{max_row}", FormulaRule(formula=["$H2>7"], stopIfTrue=True, fill=green_fill))

    # Colorear columna Usuario y Legajo ligeramente
    user_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    leg_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    for r in range(2, max_row + 1):
        ws.cell(row=r, column=6).fill = user_fill
        ws.cell(row=r, column=5).fill = leg_fill

    # Ajustar anchos de columna
    widths = [40, 15, 16, 10, 14, 20, 16, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Activar autofiltro para permitir filtrar por Usuario/Legajo en Excel
    if max_row >= 2:
        ws.auto_filter.ref = f"A1:H{max_row}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Guardar el Excel persistentemente para el usuario
    excel_path = f"excels/{usuario}.xlsx"
    wb.save(excel_path)

    headers = {"Content-Disposition": "attachment; filename=lista_final.xlsx"}
    try:
        send_email_to_user(session_id, "Lista guardada", "Su lista fue guardada en un archivo Excel.")
    except Exception:
        pass
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@app.get("/obtener_lista")
async def obtener_lista(request: Request) -> Dict[str, Any]:
    """Obtener la lista de productos guardada del usuario actual"""
    session_id = request.cookies.get("session_id")
    if not session_id:
        raise HTTPException(status_code=403, detail="No autenticado")
    
    lista = listas_por_usuario.get(session_id, [])
    return {"lista": lista}

@app.post("/modificar_producto/{codigo}")
async def modificar_producto(codigo: str, request: Request) -> Dict[str, Any]:
    data = await request.json()
    session_id = request.cookies.get("session_id")
    if not session_id or session_id not in listas_por_usuario:
        raise HTTPException(status_code=403, detail="No autenticado")
    
    lista = listas_por_usuario[session_id]
    print(f"[DEBUG] modificar_producto inicio: usuario={session_id}, lista_len={len(lista)}")
    producto_encontrado = None
    
    for p in lista:
        if p["Codigo"] == codigo:
            producto_encontrado = p
            break
    
    if not producto_encontrado:
        print(f"[DEBUG] modificar_producto: no se encontró producto {codigo}")
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    # Actualizar campos
    if "fecha_vencimiento" in data and data["fecha_vencimiento"]:
        producto_encontrado["FechaVencimiento"] = data["fecha_vencimiento"]
        estado_texto, dias_restantes = estado_vencimiento(data["fecha_vencimiento"])
        if estado_texto == "Fecha inválida":
            raise HTTPException(status_code=400, detail="Fecha de vencimiento inválida")
        producto_encontrado["Estado"] = estado_texto
        producto_encontrado["DiasRestantes"] = dias_restantes
    
    if "descripcion" in data and data["descripcion"]:
        producto_encontrado["Descripcion"] = data["descripcion"]
    
    if "ean" in data and data["ean"]:
        producto_encontrado["EAN"] = data["ean"]

    # Actualizar stock si se envía (puede ser 0)
    if "stock" in data and data["stock"] is not None:
        try:
            producto_encontrado["Stock"] = int(data["stock"])
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="Stock inválido")

    print(f"[DEBUG] modificar_producto fin: lista_len={len(lista)}")
    guardar_listas()
    return {"mensaje": "Producto modificado", "lista": lista}

@app.delete("/borrar_producto/{codigo}")
async def borrar_producto(codigo: str, request: Request) -> Dict[str, Any]:
    session_id = request.cookies.get("session_id")
    if not session_id or session_id not in listas_por_usuario:
        raise HTTPException(status_code=403, detail="No autenticado")

    # Filtrar la lista quitando el producto con ese código
    listas_por_usuario[session_id] = [
        p for p in listas_por_usuario[session_id] if p["Codigo"] != codigo
    ]

    guardar_listas()
    return {"mensaje": "Producto eliminado", "lista": listas_por_usuario[session_id]}